"""
Phase 14: Excel reporting.

generate_repository_report() builds an in-memory .xlsx workbook (never
touches disk) from a RepositoryContext -- the same object Phase 12 built
as the downstream hand-off contract. This module adds zero new
fact-gathering logic: every value written here already exists on the
RepositoryContext passed in.

Security requirement (this project's hard rule, tested explicitly): no
secret, token, or credential-bearing value may ever reach a generated
workbook. Since RepositoryContext itself never carries a token (Phase 3's
own test already proves this), the real risk this module must guard
against is nothing -- but the test suite still scans every cell of every
sheet for token-like substrings, because upstream data (a repository
description, a commit message) is untrusted content that could
coincidentally or maliciously contain something that looks like a secret,
and the export must reflect it as inert display text nonetheless.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook

from app.domain.repository_context import RepositoryContext


def generate_repository_report(context: RepositoryContext) -> bytes:
    workbook = Workbook()

    _write_summary_sheet(workbook.active, context)
    _write_languages_sheet(workbook.create_sheet("Languages"), context)
    _write_files_sheet(workbook.create_sheet("Files"), context)
    _write_test_frameworks_sheet(workbook.create_sheet("Test Frameworks"), context)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_summary_sheet(sheet, context: RepositoryContext) -> None:
    sheet.title = "Summary"
    rows = [
        ("Repository URL", context.source_url),
        ("Provider", context.provider),
        ("Owner", context.owner),
        ("Name", context.repository_name),
        ("Description", context.metadata.description if context.metadata else None),
        ("Selected branch", context.selected_branch),
        ("Default branch", context.default_branch),
        ("Commit SHA", context.commit_sha),
        ("Visibility", context.metadata.visibility if context.metadata else None),
        ("Primary language", context.metadata.primary_language if context.metadata else None),
        ("Stargazers", context.metadata.stargazers_count if context.metadata else None),
        ("Forks", context.metadata.forks_count if context.metadata else None),
        ("Analysis status", context.analysis_status.value),
        ("Exported at", datetime.now(timezone.utc).isoformat()),
    ]
    sheet.append(["Field", "Value"])
    for label, value in rows:
        sheet.append([label, value])


def _write_languages_sheet(sheet, context: RepositoryContext) -> None:
    sheet.append(["Language", "Bytes"])
    for language, byte_count in sorted(context.languages.items(), key=lambda item: -item[1]):
        sheet.append([language, byte_count])


def _write_files_sheet(sheet, context: RepositoryContext) -> None:
    sheet.append(["Relative Path", "Extension", "Category", "Language", "Size (bytes)"])
    for entry in context.file_tree:
        sheet.append([entry.relative_path, entry.extension, entry.category, entry.language, entry.size_bytes])


def _write_test_frameworks_sheet(sheet, context: RepositoryContext) -> None:
    sheet.append(["Test Framework"])
    for framework in context.test_frameworks:
        sheet.append([framework])
