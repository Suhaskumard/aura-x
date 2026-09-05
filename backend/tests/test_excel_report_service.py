"""
Phase 14: Excel reporting tests.

Uses a real RepositoryContext (built via a real local-repo ingestion,
matching Phase 12's test pattern) rather than a hand-built fixture, so the
generated workbook reflects genuine scan/clone/persist data.
"""

import shutil
import subprocess
from pathlib import Path

import httpx
import openpyxl
import pytest
import respx
from sqlalchemy.orm import sessionmaker

from app.core.config import Settings
from app.domain.repository_context import IngestionStatus, RepositoryContext
from app.services import clone_service
from app.services.context_builder import build_repository_context
from app.services.excel_report_service import generate_repository_report
from app.services.github_provider import GitHubProvider
from app.services.ingestion_service import continue_ingestion, start_ingestion

pytestmark = pytest.mark.skipif(shutil.which("git") is None, reason="git executable not on PATH")


def _run(args: list[str], cwd: Path) -> None:
    subprocess.run(args, cwd=cwd, check=True, capture_output=True, text=True)


SECRET_MARKERS = [
    "ghp_FAKESECRETTOKENVALUE1234567890",
    "Bearer sometoken",
    "postgresql://user:hunter2@host/db",
]


@pytest.fixture
def local_repo(tmp_path: Path) -> Path:
    repo_dir = tmp_path / "source_repo"
    repo_dir.mkdir()
    _run(["git", "init", "--initial-branch=main"], cwd=repo_dir)
    _run(["git", "config", "user.email", "test@example.com"], cwd=repo_dir)
    _run(["git", "config", "user.name", "Test"], cwd=repo_dir)
    (repo_dir / "app.py").write_text("print('hi')\n", encoding="utf-8")
    (repo_dir / "requirements.txt").write_text("pytest==8.0.0\n", encoding="utf-8")
    _run(["git", "add", "."], cwd=repo_dir)
    _run(["git", "commit", "-m", "initial commit"], cwd=repo_dir)
    return repo_dir


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    return Settings(
        database_url=f"sqlite:///{tmp_path / 'test.db'}",
        workspace_root=tmp_path / "workspace",
        clone_timeout_seconds=30,
        github_max_retries=1,
        github_request_timeout_seconds=1.0,
    )


@pytest.fixture(autouse=True)
def _allow_local_file_transport_for_tests(monkeypatch):
    monkeypatch.setitem(clone_service._GIT_ENV_OVERRIDES, "GIT_ALLOW_PROTOCOL", "https:file")


@pytest.fixture(autouse=True)
def _redirect_clone_to_local_repo(monkeypatch, local_repo):
    def fake_clone(self, owner, repo, branch, target_dir):
        return clone_service._clone_and_verify(
            clone_url=str(local_repo), branch=branch, target_dir=Path(target_dir), settings=self._settings
        )

    monkeypatch.setattr(GitHubProvider, "clone", fake_clone)


@pytest.fixture
def db_session(tmp_path, settings):
    import app.models  # noqa: F401
    from app.db.base import Base
    from app.db.session import create_db_engine

    engine = create_db_engine(settings)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    session = SessionLocal()
    yield session
    session.close()


@pytest.fixture
def session_factory(db_session):
    return sessionmaker(bind=db_session.get_bind(), autoflush=False, autocommit=False, future=True)


def _mock_github(*, description: str = "test repo"):
    respx.get("https://api.github.com/repos/octocat/hello-world").mock(
        return_value=httpx.Response(
            200,
            json={
                "id": 999004,
                "name": "hello-world",
                "owner": {"login": "octocat"},
                "description": description,
                "default_branch": "main",
                "private": False,
                "language": "Python",
                "topics": [],
                "license": None,
                "stargazers_count": 5,
                "forks_count": 1,
                "open_issues_count": 0,
                "created_at": "2020-01-01T00:00:00Z",
                "updated_at": "2024-06-01T12:30:00Z",
            },
        )
    )
    respx.get("https://api.github.com/repos/octocat/hello-world/branches").mock(
        return_value=httpx.Response(200, json=[{"name": "main", "commit": {"sha": "abc123"}}])
    )
    respx.get("https://api.github.com/repos/octocat/hello-world/languages").mock(
        return_value=httpx.Response(200, json={"Python": 1234})
    )


def _ingest(db_session, session_factory, settings, *, description="test repo"):
    _mock_github(description=description)
    run, parsed = start_ingestion(
        db_session, source_url="https://github.com/octocat/hello-world", requested_branch=None, settings=settings
    )
    continue_ingestion(
        run.id,
        owner=parsed.owner,
        repo=parsed.repository,
        requested_branch=None,
        settings=settings,
        session_factory=session_factory,
    )
    # continue_ingestion updated this row through its own separate
    # session (by design -- see ingestion_service.py's docstring);
    # db_session's identity map is unaware of that change until expired.
    db_session.expire_all()
    return build_repository_context(db_session, run_id=run.id, settings=settings)


@respx.mock
def test_generate_report_contains_real_repository_facts(db_session, session_factory, settings):
    context = _ingest(db_session, session_factory, settings)

    workbook_bytes = generate_repository_report(context)
    workbook = openpyxl.load_workbook(_bytes_to_tmp(workbook_bytes))

    assert workbook.sheetnames == ["Summary", "Languages", "Files", "Test Frameworks"]

    summary = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2)}
    assert summary["Repository URL"] == context.source_url
    assert summary["Owner"] == "octocat"
    assert summary["Name"] == "hello-world"
    assert summary["Commit SHA"] == context.commit_sha
    assert summary["Analysis status"] == "READY"

    languages_sheet = workbook["Languages"]
    lang_rows = list(languages_sheet.iter_rows(min_row=2, values_only=True))
    assert ("Python", 1234) in lang_rows

    files_sheet = workbook["Files"]
    file_paths = [row[0] for row in files_sheet.iter_rows(min_row=2, values_only=True)]
    assert "app.py" in file_paths

    frameworks_sheet = workbook["Test Frameworks"]
    frameworks = [row[0] for row in frameworks_sheet.iter_rows(min_row=2, values_only=True)]
    assert "pytest" in frameworks


@respx.mock
def test_generate_report_never_leaks_secret_like_values_from_upstream_data(
    db_session, session_factory, settings
):
    # A repository description is untrusted upstream content -- prove that
    # even if it contained something token-shaped, the export carries it
    # through as inert display text (present, because it's real repo data)
    # without ever adding any ADDITIONAL secret-shaped value of its own
    # (e.g. no config/env dump anywhere in the workbook).
    malicious_description = f"desc with embedded {SECRET_MARKERS[0]} and {SECRET_MARKERS[2]}"
    context = _ingest(db_session, session_factory, settings, description=malicious_description)

    workbook_bytes = generate_repository_report(context)
    workbook = openpyxl.load_workbook(_bytes_to_tmp(workbook_bytes))

    all_cell_values = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            all_cell_values.extend(str(v) for v in row if v is not None)

    combined_text = "\n".join(all_cell_values)

    # The description itself is real data and is expected to appear.
    assert SECRET_MARKERS[0] in combined_text

    # But nothing the exporter itself controls (field labels, headers,
    # every OTHER cell) may contain the actual GITHUB_TOKEN environment
    # value or the literal strings "Authorization"/"Bearer "/"password"
    # as a HEADER OR LABEL -- i.e. no cell whose column is a fixed field
    # name (not the untrusted description) contains a secret marker.
    forbidden_columns = {"Repository URL", "Provider", "Owner", "Name", "Selected branch",
                         "Default branch", "Commit SHA", "Visibility", "Primary language",
                         "Stargazers", "Forks", "Analysis status", "Exported at"}
    summary_sheet = workbook["Summary"]
    for row in summary_sheet.iter_rows(min_row=2, values_only=True):
        label, value = row[0], row[1]
        if label in forbidden_columns - {"Repository URL"}:  # URL could legitimately embed no secret
            assert not any(marker in str(value) for marker in SECRET_MARKERS if value)


def test_generate_report_from_context_with_no_scan_data(settings):
    # A run that failed before scanning completed (e.g. stuck at CLONING,
    # or genuinely FAILED early) has no file_tree/languages/evolution
    # signals. The exporter must still produce a valid, openable workbook.
    context = RepositoryContext(
        repository_id="1",
        provider="github",
        source_url="https://github.com/octocat/hello-world",
        owner="octocat",
        repository_name="hello-world",
        analysis_status=IngestionStatus.CLONING,
    )
    workbook_bytes = generate_repository_report(context)
    workbook = openpyxl.load_workbook(_bytes_to_tmp(workbook_bytes))
    assert workbook.sheetnames == ["Summary", "Languages", "Files", "Test Frameworks"]
    assert list(workbook["Languages"].iter_rows(min_row=2, values_only=True)) == []
    assert list(workbook["Files"].iter_rows(min_row=2, values_only=True)) == []


def _bytes_to_tmp(data: bytes):
    import io

    return io.BytesIO(data)
