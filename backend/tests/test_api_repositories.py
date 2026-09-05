import shutil
import subprocess
from pathlib import Path

import httpx
import pytest
import respx

from app.db.session import get_db
from app.main import app
from app.services import clone_service
from app.services.github_provider import GitHubProvider

pytestmark = pytest.mark.skipif(shutil.which("git") is None, reason="git executable not on PATH")

REPO_PAYLOAD = {
    "id": 999002,
    "name": "hello-world",
    "owner": {"login": "octocat"},
    "description": "test repo",
    "default_branch": "main",
    "private": False,
    "language": "Python",
    "topics": [],
    "license": None,
    "stargazers_count": 1,
    "forks_count": 0,
    "open_issues_count": 0,
    "created_at": "2020-01-01T00:00:00Z",
    "updated_at": "2024-06-01T12:30:00Z",
}
BRANCHES_PAYLOAD = [
    {"name": "main", "commit": {"sha": "abc123"}},
    {"name": "dev", "commit": {"sha": "def456"}},
]
LANGUAGES_PAYLOAD = {"Python": 1234}
COMMITS_PAYLOAD = [
    {
        "sha": "abc123",
        "parents": [],
        "commit": {"message": "initial", "author": {"name": "A", "email": "a@b.com", "date": "2024-01-01T00:00:00Z"}},
    }
]


def _run(args: list[str], cwd: Path) -> None:
    subprocess.run(args, cwd=cwd, check=True, capture_output=True, text=True)


@pytest.fixture
def local_repo(tmp_path: Path) -> Path:
    repo_dir = tmp_path / "source_repo"
    repo_dir.mkdir()
    _run(["git", "init", "--initial-branch=main"], cwd=repo_dir)
    _run(["git", "config", "user.email", "test@example.com"], cwd=repo_dir)
    _run(["git", "config", "user.name", "Test"], cwd=repo_dir)
    (repo_dir / "app.py").write_text("print('hi')\n", encoding="utf-8")
    _run(["git", "add", "."], cwd=repo_dir)
    _run(["git", "commit", "-m", "initial commit"], cwd=repo_dir)
    _run(["git", "branch", "dev"], cwd=repo_dir)
    return repo_dir


@pytest.fixture(autouse=True)
def _allow_local_file_transport_for_tests(monkeypatch):
    monkeypatch.setitem(clone_service._GIT_ENV_OVERRIDES, "GIT_ALLOW_PROTOCOL", "https:file")


@pytest.fixture(autouse=True)
def _redirect_clone_to_local_repo(monkeypatch, local_repo):
    def fake_clone(self, owner, repo, branch, target_dir):
        return clone_service._clone_and_verify(
            clone_url=str(local_repo), branch=branch, target_dir=Path(target_dir), settings=self._settings
        )

    monkeypatch.setattr(GitHubProvider, "clone", fake_clone)


@pytest.fixture(autouse=True)
def _use_test_workspace(monkeypatch, tmp_path):
    # Route the API's get_settings()-derived provider/clone_service calls
    # at a per-test workspace directory instead of the real one.
    from app.core.config import Settings, get_settings

    test_settings = Settings(
        workspace_root=tmp_path / "workspace",
        clone_timeout_seconds=30,
        github_max_retries=1,
        github_request_timeout_seconds=1.0,
    )
    get_settings.cache_clear()
    monkeypatch.setattr("app.core.config.get_settings", lambda: test_settings)
    monkeypatch.setattr("app.services.ingestion_service.get_settings", lambda: test_settings)
    monkeypatch.setattr("app.api.v1.repositories.get_settings", lambda: test_settings)
    yield test_settings
    get_settings.cache_clear()


def _mock_github():
    respx.get("https://api.github.com/repos/octocat/hello-world").mock(
        return_value=httpx.Response(200, json=REPO_PAYLOAD)
    )
    respx.get("https://api.github.com/repos/octocat/hello-world/branches").mock(
        return_value=httpx.Response(200, json=BRANCHES_PAYLOAD)
    )
    respx.get("https://api.github.com/repos/octocat/hello-world/languages").mock(
        return_value=httpx.Response(200, json=LANGUAGES_PAYLOAD)
    )
    respx.get("https://api.github.com/repos/octocat/hello-world/commits").mock(
        return_value=httpx.Response(200, json=COMMITS_PAYLOAD)
    )


@respx.mock
def test_create_repository_returns_202_with_in_progress_body(client_with_db):
    _mock_github()
    response = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    assert response.status_code == 202
    body = response.json()
    assert body["repository"]["owner"] == "octocat"
    # The response body is serialized at creation time, before the
    # background task runs -- it must NOT show the final outcome.
    assert body["analysis_run"]["status"] == "FETCHING_METADATA"


@respx.mock
def test_create_repository_background_task_reaches_ready_by_the_time_testclient_returns(client_with_db):
    # Starlette's TestClient runs BackgroundTasks synchronously to
    # completion before .post() returns -- so even though the response
    # BODY captured the in-progress state, the actual DB row must already
    # reflect the final outcome by now.
    _mock_github()
    response = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    run_id = response.json()["analysis_run"]["id"]

    status_resp = client_with_db.get(f"/api/v1/analysis-runs/{run_id}")
    assert status_resp.status_code == 200
    body = status_resp.json()
    assert body["status"] == "READY"
    assert body["commit_sha"] is not None
    assert body["branch_id"] is not None


def test_create_repository_invalid_url_returns_400(client_with_db):
    response = client_with_db.post("/api/v1/repositories", json={"source_url": "not a url"})
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "INVALID_REPOSITORY_URL"


@respx.mock
def test_get_repository_profile(client_with_db):
    _mock_github()
    create_resp = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    repo_id = create_resp.json()["repository"]["id"]

    response = client_with_db.get(f"/api/v1/repositories/{repo_id}")
    assert response.status_code == 200
    assert response.json()["name"] == "hello-world"


def test_get_repository_profile_404_for_unknown_id(client_with_db):
    response = client_with_db.get("/api/v1/repositories/does-not-exist")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "REPOSITORY_NOT_FOUND"


@respx.mock
def test_list_branches(client_with_db):
    _mock_github()
    create_resp = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    repo_id = create_resp.json()["repository"]["id"]

    response = client_with_db.get(f"/api/v1/repositories/{repo_id}/branches")
    assert response.status_code == 200
    names = {b["name"] for b in response.json()}
    assert names == {"main", "dev"}


@respx.mock
def test_list_commits(client_with_db):
    _mock_github()
    create_resp = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    repo_id = create_resp.json()["repository"]["id"]

    response = client_with_db.get(f"/api/v1/repositories/{repo_id}/commits")
    assert response.status_code == 200
    assert response.json()[0]["sha"] == "abc123"


@respx.mock
def test_refresh_repository_creates_new_analysis_run(client_with_db):
    _mock_github()
    create_resp = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    repo_id = create_resp.json()["repository"]["id"]
    first_run_id = create_resp.json()["analysis_run"]["id"]

    refresh_resp = client_with_db.post(f"/api/v1/repositories/{repo_id}/refresh")
    assert refresh_resp.status_code == 202
    assert refresh_resp.json()["analysis_run"]["id"] != first_run_id


@respx.mock
def test_refresh_repository_with_explicit_branch(client_with_db):
    _mock_github()
    create_resp = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    repo_id = create_resp.json()["repository"]["id"]

    refresh_resp = client_with_db.post(
        f"/api/v1/repositories/{repo_id}/refresh", json={"branch": "dev"}
    )
    assert refresh_resp.status_code == 202
    run_id = refresh_resp.json()["analysis_run"]["id"]

    status_resp = client_with_db.get(f"/api/v1/analysis-runs/{run_id}")
    assert status_resp.status_code == 200
    body = status_resp.json()
    assert body["status"] == "READY", body
    assert body["requested_branch"] == "dev"


def test_list_repositories_pagination_bounds(client_with_db):
    response = client_with_db.get("/api/v1/repositories", params={"limit": 0})
    assert response.status_code == 422  # ge=1 violated

    response = client_with_db.get("/api/v1/repositories", params={"limit": 101})
    assert response.status_code == 422  # le=100 violated

    response = client_with_db.get("/api/v1/repositories", params={"offset": -1})
    assert response.status_code == 422  # ge=0 violated


@respx.mock
def test_list_repositories_returns_created_repository(client_with_db):
    _mock_github()
    client_with_db.post("/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"})

    response = client_with_db.get("/api/v1/repositories")
    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["owner"] == "octocat"


def test_get_analysis_run_404_for_unknown_id(client_with_db):
    response = client_with_db.get("/api/v1/analysis-runs/999999")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "REPOSITORY_NOT_FOUND"


@respx.mock
def test_stuck_run_reconciliation_forces_failed(client_with_db, db_session, monkeypatch):
    # A run that's been stuck (non-terminal, stale updated_at) longer than
    # stuck_run_timeout_seconds must be force-failed the next time it's
    # polled, without ever violating the state machine. The background
    # task already completed (TestClient runs it inline) so this run is
    # actually READY -- force it back to a non-terminal state with a
    # stale updated_at to simulate a genuinely stuck run.
    from datetime import datetime, timedelta, timezone

    from app.core.config import Settings
    from app.models.analysis_run import AnalysisRun

    _mock_github()
    create_resp = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    run_id = create_resp.json()["analysis_run"]["id"]

    run = db_session.get(AnalysisRun, run_id)
    run.status = "CLONING"
    run.updated_at = datetime.now(timezone.utc) - timedelta(seconds=1000)
    db_session.commit()

    stuck_settings = Settings(stuck_run_timeout_seconds=600)
    monkeypatch.setattr("app.api.v1.repositories.get_settings", lambda: stuck_settings)

    response = client_with_db.get(f"/api/v1/analysis-runs/{run_id}")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "FAILED"
    assert body["last_error"]["code"] == "STUCK_RUN_TIMEOUT"


@respx.mock
def test_export_analysis_run_returns_valid_xlsx(client_with_db):
    import io

    import openpyxl

    _mock_github()
    create_resp = client_with_db.post(
        "/api/v1/repositories", json={"source_url": "https://github.com/octocat/hello-world"}
    )
    run_id = create_resp.json()["analysis_run"]["id"]

    response = client_with_db.get(f"/api/v1/analysis-runs/{run_id}/export.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert "octocat-hello-world" in response.headers["content-disposition"]

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Languages", "Files", "Test Frameworks"]
    summary = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2)}
    assert summary["Owner"] == "octocat"
    assert summary["Name"] == "hello-world"


def test_export_analysis_run_404_for_unknown_run(client_with_db):
    response = client_with_db.get("/api/v1/analysis-runs/999999/export.xlsx")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "REPOSITORY_NOT_FOUND"
